# -*- coding:utf-8 -*-
# created by Alex Li - 路飞学城

from openpyxl import load_workbook
import smtplib

from email.mime.text import MIMEText # 邮件正文
from email.header import Header  # 邮件头

# 加载excel文件 ，
wb = load_workbook("大唐建设集团-2022年5月工资.xlsx",data_only=True)

sheet = wb.active


# 登录
smtp_obj = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)
smtp_obj.login("nami@luffycity.com", "FfV2PjWxRn")

# 循环excel
count = 0
table_col_html = '<thead>'  # 表头
for row in sheet.iter_rows(min_row=1):
    count += 1
    if count == 1 : # first row
        for col in row:
            table_col_html += f"<th>{col.value}</th>"
        table_col_html += "</thead>"
        continue
    else:
        row_text = "<tr>" # 开始一行
        for cell in row:
            row_text += f"<td>{cell.value}</td>"
        row_text += "</tr>" # 结束一行
        name = row[2]
        staff_email = row[1].value
        print(staff_email,name.value)


    mail_body_context = f'''
        <h3>{name.value},你好:</h3>
        <p>请查收你2022-05月的工资条， 钱少了，别他妈的来找我。。。。</p>
        <table border="1px solid black"> 
          {table_col_html}
         {row_text}
       </table>
    '''
    msg_body = MIMEText(mail_body_context,"html","utf-8")

    msg_body["From"] = Header("大唐人事部", "utf-8")  # 发送者
    msg_body["To"] = Header("大唐员工", "utf-8")  # 接收者
    msg_body["Subject"] = Header("大唐建设集团2022-05月工资", "utf-8")  # 主题

    # 发邮件
    smtp_obj.sendmail("nami@luffycity.com", [staff_email,],msg_body.as_string() )
    print(f"成功发送工资条到{staff_email}- {name.value}....")