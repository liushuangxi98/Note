# -*- coding:utf-8 -*-
# created by Alex Li - 路飞学城
from openpyxl import Workbook, load_workbook
# #
# wb = Workbook() # create an excel file in RAM
# sheet = wb.active
#
# print(sheet.title)
# sheet.title = "Alex大王的姑娘们"
#
#
# sheet["B9"] = "black girl"
# sheet["C9"] = "171, 48,84 "
# sheet.append(["Rachel","170","49"])
# wb.save("excel_test.xlsx")
#
#
# # 打开已有文件
# # wb = load_workbook("excel_test.xlsx")
#

wb = load_workbook("集训营报名人数.xlsx")

print(wb.sheetnames)
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name("Sheet1")
print(sheet["B5"])
print(sheet["B5"].value)

# for cell in sheet["B5:B10"]: 获取指定列的切片数据
#     print(cell[0].value)

# for row in sheet:
#     #print(row )
#     for cell in row:
#         print(cell.value,end=",")
#
#     print()

# for row in sheet.iter_rows(min_row=4,max_row=21, max_col=5):
#     for cell in row:
#         print(cell.value,end=",")
#     print()

# 按列循环
# for col in sheet.columns:
#     for cell in col:
#         print(cell.value, end=",")
#     print()

# for col in sheet.iter_cols(min_col=3,max_col=4,min_row=3,max_row=20):
#     for cell in col:
#         print(cell.value, end=",")
#     print()

sheet["B5"] = "金角大王"

from openpyxl.styles import Font, colors, Alignment,Border,Side

myfont = Font(name="宋体",size=20,italic=True, color=colors.YELLOW)


sheet["B5"].font = myfont
sheet["B5"].alignment = Alignment(vertical="center", horizontal="center")


# 第2行行高
sheet.row_dimensions[2].height = 40
# C列列宽
sheet.column_dimensions['C'].width = 30



# border

border = Border(left=Side(border_style='medium', #dark light
                          color=colors.BLACK),
                right=Side(border_style='medium',
                          color=colors.RED),
                top=Side(border_style='medium',
                        color=colors.BLACK),
                bottom=Side(border_style='medium',
                            color=colors.BLACK),
                diagonal=Side(border_style='medium',
                              color=colors.BLACK),
                diagonal_direction=0,
                outline=Side(border_style='medium',
                            color=colors.BLACK),
                vertical=Side(border_style='medium',
                              color=colors.BLACK),
                horizontal=Side(border_style='medium',
                              color=colors.BLACK)
              )

sheet["C5"].border = border

wb.save("集训营报名人数.xlsx")