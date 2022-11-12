#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 5/11/2022 下午8:07
# @Author  : 刘双喜
# @File    : 0..py
# @Description : 添加描述
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment, fills

data = load_workbook('test_xlsx.xlsx')
sheet = data['Sheet1']
myfont = Font(name="宋体", size=20, italic=True, color=colors.BLUE)
my_fill = fills.PatternFill('solid', fgColor='FFC0CB')
for row in sheet:
    for cell in row:
        # cell.value = 55
        # cell.font = myfont
        cell.fill = my_fill
data.save('test_xlsx.xlsx')
