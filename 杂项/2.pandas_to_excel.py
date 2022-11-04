import pandas as pd
import random
import numpy as np
from openpyxl import load_workbook

data = np.random.randint(10, 30, 100).reshape(10, 10)
pd_data_3 = pd.DataFrame(data, index=pd.date_range('2022-03-01', periods=10).date,
                         columns=[f'{i}时' for i in range(10)])
# pd_data_3.to_excel('test.xlsx', sheet_name='3月', header=True, index=True, na_rep='NA', startcol=0, startrow=0)
# header:是否写入列索引
# index:是否写入行索引
# na_rep:缺失值写入为
# startcol:开始列
# startrow:开始行

with pd.ExcelWriter('test.xlsx') as writer:
    pd_data_3.to_excel(writer, sheet_name='Sheet1', index=False)
    pd_data_3.to_excel(writer, sheet_name='Sheet2', index=False)

with pd.ExcelWriter('test.xlsx', mode='a') as writer:
    pd_data_3.to_excel(writer, sheet_name='Sheet3', index=False)

# # #
book = load_workbook('test.xlsx')
pd_data_2 = pd.DataFrame([[99, 99, 99]])
with pd.ExcelWriter('test.xlsx') as writer:
    writer.book = book  # 读取excel
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)  # 复制excel的所有表
    pd_data_2.to_excel(writer, sheet_name='Sheet1', header=False, index=False,startrow=4)
    pd_data_2.to_excel(writer, sheet_name='Sheet2', header=False, index=False,startrow=4)
#
# book = load_workbook('excel1.xlsx')
# pd_data_3 = pd.DataFrame({'姓名': ['lsx', 'mc', 'ljb_sb'],})
# with pd.ExcelWriter('excel1.xlsx') as writer:
#     writer.book = book
#     writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#     pd_data_3.to_excel(writer, sheet_name='Sheet2', index=False, header=False,startrow=3, startcol=0)
